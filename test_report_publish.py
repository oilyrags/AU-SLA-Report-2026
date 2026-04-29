#!/usr/bin/env python3
import contextlib
import io
import json
import pathlib
import tempfile
import unittest

import pandas as pd
from openpyxl import load_workbook

import report_publish
import sla_report_end2end


FIXTURE_PATH = pathlib.Path(__file__).with_name("tests").joinpath("fixtures", "report_input.json")


def load_fixture_frames():
    with FIXTURE_PATH.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    all_df = pd.DataFrame(payload["all_df"])
    cycles_df = pd.DataFrame(payload["cycles_df"])
    _, in_scope, exceptions_non_type, exceptions_rejected, exceptions_feature = sla_report_end2end.classify_scope(all_df)
    return all_df, in_scope, exceptions_non_type, exceptions_rejected, exceptions_feature, cycles_df


class PublishContractTests(unittest.TestCase):
    def test_publish_sheet_payloads_invokes_adapter_for_each_tab_in_order(self):
        calls = []

        class FakeAdapter:
            def write_tab(self, tab_name, dataframe):
                calls.append((tab_name, list(dataframe.columns)))

        payloads = {
            "Dashboard": pd.DataFrame([{"metric": "FRT Attainment %", "value": 1.0}]),
            "Refresh Control": pd.DataFrame([{"last_refresh_status": "Succeeded"}]),
        }

        report_publish.publish_sheet_payloads(FakeAdapter(), payloads)

        self.assertEqual(calls[0][0], "Dashboard")
        self.assertEqual(calls[1][0], "Refresh Control")
        self.assertEqual(calls[0][1], ["metric", "value"])
        self.assertEqual(calls[1][1], ["last_refresh_status"])

    def test_dashboard_payload_ignores_rows_without_sla_coverage(self):
        in_scope = pd.DataFrame(
            [
                {
                    "frt_has_sla": True,
                    "frt_pending": False,
                    "frt_breached": False,
                    "res_has_sla": True,
                    "res_pending": False,
                    "res_breached": False,
                },
                {
                    "frt_has_sla": True,
                    "frt_pending": False,
                    "frt_breached": True,
                    "res_has_sla": True,
                    "res_pending": False,
                    "res_breached": True,
                },
                {
                    "frt_has_sla": False,
                    "frt_pending": False,
                    "frt_breached": False,
                    "res_has_sla": False,
                    "res_pending": False,
                    "res_breached": False,
                },
            ]
        )

        dashboard = report_publish.build_dashboard_payload(in_scope).set_index("metric")["value"].to_dict()

        self.assertEqual(dashboard["Tickets In Scope"], 3)
        self.assertEqual(dashboard["FRT Attainment %"], 0.5)
        self.assertEqual(dashboard["Resolution Attainment %"], 0.5)

    def test_build_sheet_payloads_returns_expected_tabs(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()

        payloads = report_publish.build_sheet_payloads(
            all_df=all_df,
            in_scope=in_scope,
            exceptions_non_type=non_type,
            exceptions_rejected=rejected,
            exceptions_feature=feature,
            cycles_df=cycles_df,
            base_url="https://example.atlassian.net",
            generated_at_iso="2026-04-24T12:00:00+00:00",
            backup_reference="backup-20260424-120000",
        )

        self.assertEqual(
            list(payloads.keys()),
            [
                "Dashboard",
                "Summary",
                "Teams",
                "Breaches",
                "Trends",
                "Exceptions",
                "Methodology",
                "Raw Data",
                "Refresh Control",
            ],
        )
        refresh_control = payloads["Refresh Control"]
        self.assertEqual(refresh_control.columns.tolist(), ["field", "value"])

        refresh_fields = refresh_control.set_index("field")["value"].to_dict()
        self.assertEqual(refresh_fields["last_refresh_status"], "Succeeded")
        self.assertEqual(refresh_fields["generated_at"], "2026-04-24T12:00:00+00:00")
        self.assertEqual(refresh_fields["backup_reference"], "backup-20260424-120000")
        self.assertEqual(refresh_fields["source_base_url"], "https://example.atlassian.net")

    def test_refresh_control_payload_preserves_request_metadata(self):
        refresh_control = report_publish.build_refresh_control_payload(
            status="Succeeded",
            generated_at_iso="2026-04-24T12:00:00+00:00",
            backup_reference="backup-20260424-120000",
            message="Refresh completed successfully.",
            request_metadata={
                "requested_at": "2026-04-24 13:00:00 CEST",
                "requested_by": "operator@example.com",
            },
        )

        fields = refresh_control.set_index("field")["value"].to_dict()
        self.assertEqual(fields["requested_at"], "2026-04-24 13:00:00 CEST")
        self.assertEqual(fields["requested_by"], "operator@example.com")

    def test_dashboard_payload_contains_exec_kpis(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()

        payloads = report_publish.build_sheet_payloads(
            all_df=all_df,
            in_scope=in_scope,
            exceptions_non_type=non_type,
            exceptions_rejected=rejected,
            exceptions_feature=feature,
            cycles_df=cycles_df,
            base_url="https://example.atlassian.net",
            generated_at_iso="2026-04-24T12:00:00+00:00",
            backup_reference="backup-20260424-120000",
        )

        dashboard = payloads["Dashboard"]
        self.assertIn("metric", dashboard.columns.tolist())
        self.assertIn("value", dashboard.columns.tolist())
        self.assertTrue((dashboard["metric"] == "FRT Attainment %").any())
        self.assertTrue((dashboard["metric"] == "Resolution Attainment %").any())

    def test_exceptions_payload_is_deduplicated_and_includes_feature_row(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()

        payloads = report_publish.build_sheet_payloads(
            all_df=all_df,
            in_scope=in_scope,
            exceptions_non_type=non_type,
            exceptions_rejected=rejected,
            exceptions_feature=feature,
            cycles_df=cycles_df,
            base_url="https://example.atlassian.net",
            generated_at_iso="2026-04-24T12:00:00+00:00",
            backup_reference="backup-20260424-120000",
        )

        exceptions = payloads["Exceptions"]
        self.assertIn("ASD-1004", set(rejected["key"].tolist()))
        self.assertIn("ASD-1004", set(feature["key"].tolist()))
        self.assertEqual(len(exceptions), 2)
        self.assertEqual(exceptions["key"].nunique(), len(exceptions))
        self.assertEqual(set(exceptions["key"].tolist()), {"ASD-1003", "ASD-1004"})
        self.assertTrue((exceptions["key"] == "ASD-1004").any())

    def test_workbook_exceptions_sheet_keeps_feature_request_out_of_rejected_table(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()
        in_scope = in_scope.copy()
        in_scope["business_to_wall_clock_ratio"] = pd.NA
        in_scope["url"] = "https://example.atlassian.net/browse/ASD-1004"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            sla_report_end2end.build_workbook(
                all_df=all_df,
                in_scope=in_scope,
                exceptions_non_type=non_type,
                exceptions_rejected=rejected,
                exceptions_feature=feature,
                cycles_df=cycles_df,
                raw_json_path="/tmp/report_input.json",
                output_path=output_path,
                base_url="https://example.atlassian.net",
            )

            workbook = load_workbook(output_path)
            values = [
                cell
                for row in workbook["Exceptions"].iter_rows(values_only=True)
                for cell in row
                if cell is not None
            ]

            self.assertEqual(values.count("ASD-1004"), 1)
            self.assertIn("ASD-1003", values)
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)

    def test_workbook_warns_when_feature_exceptions_lack_intent_labels(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()
        in_scope = in_scope.copy()
        in_scope["business_to_wall_clock_ratio"] = pd.NA
        in_scope["url"] = "https://example.atlassian.net/browse/ASD-1004"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            stderr = io.StringIO()
            with contextlib.redirect_stderr(stderr):
                sla_report_end2end.build_workbook(
                    all_df=all_df,
                    in_scope=in_scope,
                    exceptions_non_type=non_type,
                    exceptions_rejected=rejected,
                    exceptions_feature=feature.drop(columns=[col for col in feature.columns if col == "intent_label"]),
                    cycles_df=cycles_df,
                    raw_json_path="/tmp/report_input.json",
                    output_path=output_path,
                    base_url="https://example.atlassian.net",
                )

            self.assertIn("exceptions_feature missing intent label column", stderr.getvalue())
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)

    def test_workbook_executive_dashboard_uses_modern_sheets_style_layout(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()
        in_scope = in_scope.copy()
        in_scope["business_to_wall_clock_ratio"] = pd.NA
        in_scope["url"] = "https://example.atlassian.net/browse/ASD-1004"
        in_scope.loc[in_scope.index[0], "res_pending"] = True
        in_scope.loc[in_scope.index[0], "res_breached"] = True

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            sla_report_end2end.build_workbook(
                all_df=all_df,
                in_scope=in_scope,
                exceptions_non_type=non_type,
                exceptions_rejected=rejected,
                exceptions_feature=feature,
                cycles_df=cycles_df,
                raw_json_path="/tmp/report_input.json",
                output_path=output_path,
                base_url="https://example.atlassian.net",
            )

            workbook = load_workbook(output_path)
            dashboard = workbook["Executive Dashboard"]

            self.assertFalse(dashboard.sheet_view.showGridLines)
            self.assertIn("A1:L1", {str(rng) for rng in dashboard.merged_cells.ranges})
            self.assertEqual(dashboard["A1"].value, "SLA Executive Dashboard")
            self.assertEqual(dashboard["A1"].fill.fgColor.rgb, "FF153E4D")

            self.assertEqual(dashboard["A4"].value, "FRT Attainment")
            self.assertEqual(dashboard["D4"].value, "Resolution Attainment")
            self.assertEqual(dashboard["G4"].value, "Tickets In Scope")
            self.assertEqual(dashboard["J4"].value, "Open Risk")
            self.assertIsNotNone(dashboard["A5"].value)
            self.assertIsNotNone(dashboard["D5"].value)
            self.assertEqual(dashboard["J5"].value, 1)
            self.assertNotIn("%", dashboard["J5"].number_format)

            self.assertEqual(dashboard["A9"].value, "Executive Summary")
            self.assertEqual(dashboard["H9"].value, "Priority Health")
            self.assertEqual(dashboard["A18"].value, "Recommended Actions")
            self.assertEqual(dashboard["H18"].value, "Top Risks")
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)

    def test_workbook_supporting_tabs_use_modern_readable_layout(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()
        in_scope = in_scope.copy()
        in_scope["business_to_wall_clock_ratio"] = pd.NA
        in_scope["url"] = "https://example.atlassian.net/browse/ASD-1004"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            sla_report_end2end.build_workbook(
                all_df=all_df,
                in_scope=in_scope,
                exceptions_non_type=non_type,
                exceptions_rejected=rejected,
                exceptions_feature=feature,
                cycles_df=cycles_df,
                raw_json_path="/tmp/report_input.json",
                output_path=output_path,
                base_url="https://example.atlassian.net",
            )

            workbook = load_workbook(output_path)
            expected_titles = {
                "SLA Summary": "SLA Summary",
                "Team Analysis": "Team Analysis",
                "Breach Detail": "Breach Detail",
                "Trends & Patterns": "Trends & Patterns",
                "Narrative & Recommendations": "Narrative & Recommendations",
                "Exceptions": "Exceptions",
                "Methodology": "Methodology",
            }

            for sheet_name, title in expected_titles.items():
                sheet = workbook[sheet_name]
                self.assertFalse(sheet.sheet_view.showGridLines, sheet_name)
                self.assertEqual(sheet["A1"].value, title)
                self.assertEqual(sheet["A1"].fill.fgColor.rgb, "FF153E4D")
                self.assertEqual(sheet["A1"].font.color.rgb, "FFFFFFFF")

            self.assertEqual(workbook["Team Analysis"].freeze_panes, "A4")
            self.assertEqual(workbook["Breach Detail"].freeze_panes, "A4")
            self.assertEqual(workbook["Trends & Patterns"].freeze_panes, "A4")
            self.assertEqual(workbook["Exceptions"].freeze_panes, "A4")
            self.assertIsNotNone(workbook["Team Analysis"].auto_filter.ref)
            self.assertIsNotNone(workbook["Breach Detail"].auto_filter.ref)
            self.assertIsNotNone(workbook["Exceptions"].auto_filter.ref)
            self.assertEqual(workbook["Narrative & Recommendations"]["A3"].value, "EXECUTIVE SUMMARY")
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)

    def test_workbook_outputs_inclusion_logic_for_all_tickets(self):
        all_df, in_scope, non_type, rejected, feature, cycles_df = load_fixture_frames()
        in_scope = in_scope.copy()
        in_scope["business_to_wall_clock_ratio"] = pd.NA
        in_scope["url"] = "https://example.atlassian.net/browse/ASD-1004"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            sla_report_end2end.build_workbook(
                all_df=all_df,
                in_scope=in_scope,
                exceptions_non_type=non_type,
                exceptions_rejected=rejected,
                exceptions_feature=feature,
                cycles_df=cycles_df,
                raw_json_path="/tmp/report_input.json",
                output_path=output_path,
                base_url="https://example.atlassian.net",
            )

            workbook = load_workbook(output_path)
            sheet = workbook["Inclusion Logic"]
            headers = [sheet.cell(3, col).value for col in range(1, sheet.max_column + 1)]
            keys = {
                sheet.cell(row, 1).value
                for row in range(4, sheet.max_row + 1)
                if sheet.cell(row, 1).value
            }

            self.assertEqual(sheet["A1"].value, "Inclusion Logic")
            self.assertIn("Scope Decision", headers)
            self.assertIn("Scope Reason", headers)
            self.assertEqual(keys, set(all_df["key"].tolist()))
            self.assertIsNotNone(sheet.auto_filter.ref)
            self.assertFalse(sheet.sheet_view.showGridLines)
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
