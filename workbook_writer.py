#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
import time
from datetime import datetime, timezone
from statistics import median
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_publish import build_summary_block
from sla_rules import (
    BUG_CUES,
    BUG_TYPE,
    DEFAULT_PROJECT_KEY,
    DEFAULT_YEAR,
    EMAIL_REQUEST_TYPE,
    FEATURE_REQUEST_CUES,
    IN_SCOPE_ISSUE_TYPES,
    ISSUE_TYPE,
    MANUAL_DISGUISED_FEATURE_KEYS,
    MANUAL_INCLUDE_BUG_KEYS,
    NOT_ACCEPTED_RESOLUTIONS,
    Q1_TARGET_MODE,
    SLA_FRT_NAMES,
    SLA_RES_NAMES,
    TEAM_FIELD_ID,
    TEAM_FIELD_NAME,
    canonical_priority,
    created_week_start,
    extract_name,
    get_targets,
    hours_from_millis,
    normalize_issue_type,
    parse_jira_dt,
    pctl,
    percent,
    quarter_of,
    safe_get,
)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def freeze_and_filter(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def find_sla_entry(values: List[Dict[str, Any]], names: set[str]) -> Optional[Dict[str, Any]]:
    for entry in values:
        name = str(entry.get("name", "")).strip().lower()
        if name in names:
            return entry
    return None


def summarize_sla(entry: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    out = {
        "has_sla": False,
        "pending": False,
        "breached": None,
        "elapsed_millis": None,
        "goal_millis": None,
        "remaining_millis": None,
        "completed_cycles_count": 0,
        "has_ongoing_cycle": False,
        "cycle_count_total": 0,
        "latest_start": None,
        "latest_stop": None,
        "paused": None,
        "within_calendar_hours": None,
        "raw": entry,
    }
    if not entry:
        return out

    completed = entry.get("completedCycles") or []
    ongoing = entry.get("ongoingCycle")

    out["has_sla"] = True
    out["completed_cycles_count"] = len(completed)
    out["has_ongoing_cycle"] = ongoing is not None
    out["cycle_count_total"] = len(completed) + (1 if ongoing else 0)

    if completed:
        c = completed[-1]
        out["pending"] = False
        out["breached"] = c.get("breached")
        out["elapsed_millis"] = safe_get(c, "elapsedTime", "millis")
        out["goal_millis"] = safe_get(c, "goalDuration", "millis")
        out["remaining_millis"] = safe_get(c, "remainingTime", "millis")
        out["latest_start"] = safe_get(c, "startTime", "jira") or safe_get(c, "startTime", "iso8601")
        out["latest_stop"] = safe_get(c, "stopTime", "jira") or safe_get(c, "stopTime", "iso8601")
        return out

    if ongoing:
        elapsed = safe_get(ongoing, "elapsedTime", "millis")
        goal = safe_get(ongoing, "goalDuration", "millis")
        breached = ongoing.get("breached")
        if breached is None and elapsed is not None and goal is not None:
            breached = elapsed > goal
        out["pending"] = True
        out["breached"] = breached
        out["elapsed_millis"] = elapsed
        out["goal_millis"] = goal
        out["remaining_millis"] = safe_get(ongoing, "remainingTime", "millis")
        out["latest_start"] = safe_get(ongoing, "startTime", "jira") or safe_get(ongoing, "startTime", "iso8601")
        out["paused"] = ongoing.get("paused")
        out["within_calendar_hours"] = ongoing.get("withinCalendarHours")
        return out

    out["pending"] = True
    return out


def collect_issue_data(
    client: Any,
    sleep_seconds: float,
    *,
    project_key: str = DEFAULT_PROJECT_KEY,
    year: int = DEFAULT_YEAR,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    year_start = f"{year}-01-01"
    next_year_start = f"{year + 1}-01-01"
    jql = f'project = {project_key} AND created >= "{year_start}" AND created < "{next_year_start}" ORDER BY created ASC'

    fields = [
        "summary",
        "issuetype",
        "status",
        "resolution",
        "priority",
        "created",
        "resolutiondate",
        "assignee",
        "reporter",
        "components",
        "labels",
        "project",
        TEAM_FIELD_ID,
    ]

    issues = client.search_issues(jql, fields)
    if not issues:
        raise RuntimeError("Query returned 0 issues. Check project key / permissions.")

    merged_rows: List[Dict[str, Any]] = []
    cycle_rows: List[Dict[str, Any]] = []
    raw_records: List[Dict[str, Any]] = []

    for i, issue in enumerate(issues, start=1):
        key = issue["key"]
        fields = issue["fields"]

        sla_payload = client.get_sla_cycles(key)
        changelog_payload = client.get_issue_changelog(key)

        sla_values = sla_payload.get("values", []) or []
        frt_entry = find_sla_entry(sla_values, SLA_FRT_NAMES)
        res_entry = find_sla_entry(sla_values, SLA_RES_NAMES)

        frt = summarize_sla(frt_entry)
        res = summarize_sla(res_entry)

        created = fields.get("created")
        resolutiondate = fields.get("resolutiondate")
        created_dt = parse_jira_dt(created)
        resolved_dt = parse_jira_dt(resolutiondate)
        wall_clock_hours = None
        if created_dt and resolved_dt:
            wall_clock_hours = (resolved_dt - created_dt).total_seconds() / 3600.0

        pri = canonical_priority(extract_name(fields.get("priority")))
        quarter = quarter_of(created)
        frt_target, res_target = get_targets(pri, quarter)

        issue_type = extract_name(fields.get("issuetype"))
        resolution_name = extract_name(fields.get("resolution"))
        domain = extract_name(fields.get(TEAM_FIELD_ID))
        components = extract_name(fields.get("components"))
        team = domain or components or extract_name(fields.get("assignee")) or "Unassigned"

        priority_changed = False
        downgraded_close_to_deadline = False
        history_count = 0

        histories = changelog_payload.get("values", []) or changelog_payload.get("histories", []) or []
        history_count = len(histories)
        for hist in histories:
            for item in hist.get("items", []) or []:
                if item.get("field") == "priority":
                    priority_changed = True
                    from_p = item.get("fromString")
                    to_p = item.get("toString")
                    rank = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
                    if rank.get(str(from_p), 0) > rank.get(str(to_p), 0):
                        downgraded_close_to_deadline = True

        merged_rows.append(
            {
                "key": key,
                "summary": fields.get("summary"),
                "issue_id": issue.get("id"),
                "project_key": safe_get(fields, "project", "key"),
                "project_name": safe_get(fields, "project", "name"),
                "issuetype": issue_type,
                "status": extract_name(fields.get("status")),
                "resolution": resolution_name,
                "priority_raw": extract_name(fields.get("priority")),
                "priority": pri,
                "created": created,
                "resolutiondate": resolutiondate,
                "assignee": extract_name(fields.get("assignee")),
                "reporter": extract_name(fields.get("reporter")),
                "components": components,
                "labels": ", ".join(fields.get("labels", []) or []),
                "domain": domain,
                "team": team,
                "quarter": quarter,
                "week_start": created_week_start(created),
                "created_day_of_week": created_dt.strftime("%A") if created_dt else None,
                "created_hour": created_dt.hour if created_dt else None,
                "url": f"{client.config.base_url.rstrip('/')}/browse/{key}",
                "frt_has_sla": frt["has_sla"],
                "frt_pending": frt["pending"],
                "frt_breached": frt["breached"],
                "frt_elapsed_millis": frt["elapsed_millis"],
                "frt_goal_millis": frt["goal_millis"],
                "frt_remaining_millis": frt["remaining_millis"],
                "frt_elapsed_hours": hours_from_millis(frt["elapsed_millis"]),
                "frt_goal_hours": hours_from_millis(frt["goal_millis"]),
                "frt_completed_cycles_count": frt["completed_cycles_count"],
                "frt_total_cycles": frt["cycle_count_total"],
                "res_has_sla": res["has_sla"],
                "res_pending": res["pending"],
                "res_breached": res["breached"],
                "res_elapsed_millis": res["elapsed_millis"],
                "res_goal_millis": res["goal_millis"],
                "res_remaining_millis": res["remaining_millis"],
                "res_elapsed_hours": hours_from_millis(res["elapsed_millis"]),
                "res_goal_hours": hours_from_millis(res["goal_millis"]),
                "res_completed_cycles_count": res["completed_cycles_count"],
                "res_total_cycles": res["cycle_count_total"],
                "wall_clock_hours": wall_clock_hours,
                "business_to_wall_clock_ratio": (
                    (hours_from_millis(res["elapsed_millis"]) / wall_clock_hours)
                    if hours_from_millis(res["elapsed_millis"]) is not None and wall_clock_hours and wall_clock_hours > 0
                    else None
                ),
                "priority_changed": priority_changed,
                "priority_downgrade_evidence": downgraded_close_to_deadline,
                "changelog_event_count": history_count,
            }
        )

        for sla_entry in sla_values:
            sla_name = sla_entry.get("name")
            for idx, cycle in enumerate(sla_entry.get("completedCycles", []) or [], start=1):
                cycle_rows.append(
                    {
                        "key": key,
                        "sla_name": sla_name,
                        "cycle_kind": "completed",
                        "cycle_index": idx,
                        "breached": cycle.get("breached"),
                        "start_time": safe_get(cycle, "startTime", "jira") or safe_get(cycle, "startTime", "iso8601"),
                        "stop_time": safe_get(cycle, "stopTime", "jira") or safe_get(cycle, "stopTime", "iso8601"),
                        "elapsed_millis": safe_get(cycle, "elapsedTime", "millis"),
                        "goal_millis": safe_get(cycle, "goalDuration", "millis"),
                        "remaining_millis": safe_get(cycle, "remainingTime", "millis"),
                        "breach_time": safe_get(cycle, "breachTime", "jira") or safe_get(cycle, "breachTime", "iso8601"),
                    }
                )
            ongoing = sla_entry.get("ongoingCycle")
            if ongoing:
                cycle_rows.append(
                    {
                        "key": key,
                        "sla_name": sla_name,
                        "cycle_kind": "ongoing",
                        "cycle_index": 1,
                        "breached": ongoing.get("breached"),
                        "start_time": safe_get(ongoing, "startTime", "jira") or safe_get(ongoing, "startTime", "iso8601"),
                        "stop_time": None,
                        "elapsed_millis": safe_get(ongoing, "elapsedTime", "millis"),
                        "goal_millis": safe_get(ongoing, "goalDuration", "millis"),
                        "remaining_millis": safe_get(ongoing, "remainingTime", "millis"),
                        "breach_time": safe_get(ongoing, "breachTime", "jira") or safe_get(ongoing, "breachTime", "iso8601"),
                    }
                )

        raw_records.append(
            {
                "key": key,
                "issue": issue,
                "sla": sla_payload,
                "changelog": changelog_payload,
            }
        )

        if i % 25 == 0 or i == len(issues):
            print(f"[info] enriched {i}/{len(issues)} issues", file=sys.stderr)
        time.sleep(max(sleep_seconds, 0.0))

    raw_json = {
        "metadata": {
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "project": project_key,
            "year": year,
            "team_field": {"id": TEAM_FIELD_ID, "name": TEAM_FIELD_NAME},
            "bug_mapping": BUG_TYPE,
            "issue_mapping": ISSUE_TYPE,
            "email_request_mapping": EMAIL_REQUEST_TYPE,
            "in_scope_issue_types": sorted(IN_SCOPE_ISSUE_TYPES),
            "not_accepted_resolutions": sorted(NOT_ACCEPTED_RESOLUTIONS),
            "q1_target_assumption": Q1_TARGET_MODE,
            "jql_all": jql,
        },
        "records": raw_records,
    }

    return pd.DataFrame(merged_rows), pd.DataFrame(cycle_rows), raw_json


def classify_scope(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = df.copy()

    def _text_blob(row: pd.Series) -> str:
        return " ".join(
            str(row.get(col) or "")
            for col in ("summary", "labels", "team", "domain", "components")
        ).lower()

    def _is_disguised_feature_request(row: pd.Series) -> bool:
        key = str(row.get("key") or "")
        if key in MANUAL_DISGUISED_FEATURE_KEYS:
            return True

        text = _text_blob(row)
        has_retool_signal = "retool" in text
        has_request_signal = any(cue in text for cue in FEATURE_REQUEST_CUES)
        has_bug_signal = any(cue in text for cue in BUG_CUES)
        return has_retool_signal and has_request_signal and not has_bug_signal

    in_scope_issue_types = {normalize_issue_type(issue_type) for issue_type in IN_SCOPE_ISSUE_TYPES}

    df["is_manual_bug_include"] = df["key"].isin(MANUAL_INCLUDE_BUG_KEYS)
    df["is_in_scope_issue_type"] = (
        df["issuetype"].map(normalize_issue_type).isin(in_scope_issue_types)
        | df["is_manual_bug_include"]
    )
    df["is_bug_issue_type"] = df["is_in_scope_issue_type"]
    df["is_disguised_feature_request"] = df.apply(_is_disguised_feature_request, axis=1)
    df["is_not_accepted"] = df["resolution"].isin(NOT_ACCEPTED_RESOLUTIONS)
    df["is_in_scope"] = df["is_in_scope_issue_type"] & (~df["is_not_accepted"]) & (~df["is_disguised_feature_request"])
    df["scope_decision"] = df["is_in_scope"].map(lambda is_in_scope: "Included" if is_in_scope else "Excluded")

    def _scope_reason(row: pd.Series) -> str:
        reasons = []
        if not bool(row.get("is_in_scope_issue_type")):
            issue_type = row.get("issuetype") or "blank"
            reasons.append(f"issue type '{issue_type}' is outside SLA scope")
        if bool(row.get("is_not_accepted")):
            resolution = row.get("resolution") or "blank"
            reasons.append(f"resolution '{resolution}' is excluded")
        if bool(row.get("is_disguised_feature_request")):
            reasons.append("ticket matches disguised feature request rules")

        if reasons:
            return "Excluded: " + "; ".join(reasons) + "."

        if bool(row.get("is_manual_bug_include")):
            return "Included: manual stakeholder include and no exclusion rule matched."

        return "Included: issue type is in scope and no exclusion rule matched."

    df["scope_reason"] = df.apply(_scope_reason, axis=1)

    in_scope = df[df["is_in_scope"]].copy()
    exceptions_non_type = df[(~df["is_in_scope_issue_type"]) & (~df["is_disguised_feature_request"])].copy()
    exceptions_rejected = df[df["is_not_accepted"]].copy()
    exceptions_all = pd.concat([exceptions_non_type, exceptions_rejected], ignore_index=True).drop_duplicates(subset=["key"])
    exceptions_feature = df[df["is_disguised_feature_request"]].copy()

    return df, in_scope, exceptions_non_type, exceptions_rejected, exceptions_feature


def narrative_from_data(all_df, in_scope, exceptions_non_type, exceptions_rejected, cycles_df):
    lines = []

    if len(in_scope) == 0:
        lines.append("EXECUTIVE SUMMARY")
        lines.append("No in-scope tickets were found for this report period.")
        lines.append("")
        lines.append("KEY DRIVERS OF PERFORMANCE")
        lines.append("- All pulled tickets were excluded by scope or resolution rules.")
        lines.append("")
        lines.append("PROCESS WEAKNESSES")
        lines.append("- No in-scope ticket population is available for SLA process analysis.")
        lines.append("")
        lines.append("RECOMMENDED ACTIONS")
        lines.append("- Review excluded tickets to confirm the scope rules match reporting intent.")
        lines.append("")
        lines.append("HIDDEN INSIGHTS")
        lines.append("- No in-scope breach or cycle data is available.")
        return "\n".join(lines)

    # --- Core metrics ---
    frt_pop = in_scope[in_scope["frt_pending"] == False]
    res_pop = in_scope[in_scope["res_pending"] == False]

    frt_pct = float((frt_pop["frt_breached"] == False).mean()) if len(frt_pop) else 0
    res_pct = float((res_pop["res_breached"] == False).mean()) if len(res_pop) else 0

    both = in_scope[
        (in_scope["frt_pending"] == False)
        & (in_scope["res_pending"] == False)
        & (in_scope["frt_breached"] == False)
        & (in_scope["res_breached"] == False)
    ]
    both_pct = len(both) / len(in_scope) if len(in_scope) else 0

    # Targets
    blended_frt_target = in_scope.apply(lambda r: get_targets(r["priority"], r["quarter"])[0], axis=1).dropna().mean()
    blended_res_target = in_scope.apply(lambda r: get_targets(r["priority"], r["quarter"])[1], axis=1).dropna().mean()

    # --- Headline ---
    lines.append("EXECUTIVE SUMMARY")
    lines.append(
        f"Service performance is {'on track' if frt_pct >= blended_frt_target and res_pct >= blended_res_target else 'below target'} YTD. "
        f"First Response attainment is {frt_pct:.1%} (target {blended_frt_target:.1%}), "
        f"while Resolution attainment is {res_pct:.1%} (target {blended_res_target:.1%}). "
        f"{both_pct:.1%} of tickets meet both SLAs."
    )

    # --- Key drivers ---
    lines.append("")
    lines.append("KEY DRIVERS OF PERFORMANCE")

    worst_priority = (
        in_scope.groupby("priority")["res_breached"]
        .mean()
        .sort_values(ascending=False)
        .index[0]
    )
    lines.append(f"- Resolution breaches are concentrated in {worst_priority} tickets.")

    out_of_hours = in_scope[(in_scope["created_hour"] < 8) | (in_scope["created_hour"] >= 17)]
    if len(out_of_hours):
        ooh_rate = float((out_of_hours["frt_breached"] == True).mean())
        lines.append(f"- Out-of-hours ticket intake shows elevated FRT breach risk ({ooh_rate:.1%}).")

    multi_cycles = in_scope[(in_scope["frt_total_cycles"] > 1) | (in_scope["res_total_cycles"] > 1)]
    if len(multi_cycles):
        lines.append(f"- {len(multi_cycles)} tickets show multiple SLA cycles, indicating potential workflow inefficiency or excessive waiting states.")

    rej_rate = len(exceptions_rejected) / len(all_df) if len(all_df) else 0
    lines.append(f"- {rej_rate:.1%} of tickets are rejected or not actioned, suggesting upstream triage inefficiency.")

    # --- Operational weaknesses ---
    lines.append("")
    lines.append("PROCESS WEAKNESSES")

    unassigned = in_scope["assignee"].isna().sum()
    if unassigned:
        lines.append(f"- {unassigned} tickets lack clear ownership.")

    stuck = in_scope[(in_scope["res_pending"] == True) & (in_scope["res_breached"] == True)]
    if len(stuck):
        lines.append(f"- {len(stuck)} active tickets are already beyond SLA, indicating escalation gaps.")

    prio_changes = in_scope[in_scope["priority_downgrade_evidence"] == True]
    if len(prio_changes):
        lines.append(f"- Priority downgrades detected in {len(prio_changes)} tickets, which may mask SLA risk.")

    # --- Recommendations ---
    lines.append("")
    lines.append("RECOMMENDED ACTIONS")

    lines.append("- Enforce strict triage criteria to reduce non-actionable ticket volume.")
    lines.append("- Introduce morning queue sweep to eliminate overnight FRT breaches.")
    lines.append("- Monitor tickets with repeated SLA cycles to prevent workflow abuse.")
    lines.append("- Require justification for any priority downgrade.")
    lines.append("- Focus resolution performance improvement on lowest-performing priority tier.")

    # --- Hidden insights ---
    lines.append("")
    lines.append("HIDDEN INSIGHTS")

    breached = in_scope[in_scope["res_breached"] == True]
    if len(breached):
        top5 = breached.sort_values("res_elapsed_hours", ascending=False).head(5)
        share = top5["res_elapsed_hours"].sum() / max(breached["res_elapsed_hours"].sum(), 1)
        lines.append(f"- Top 5 tickets account for {share:.1%} of total breach time.")

    ratio = in_scope["business_to_wall_clock_ratio"].dropna()
    if len(ratio):
        lines.append(f"- Median business-to-wall-clock ratio is {ratio.median():.2f}, indicating significant paused time in workflow.")

    return "\n".join(lines)


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, index: bool = False) -> None:
    headers = list(df.columns)
    for c, h in enumerate(headers, start=start_col):
        ws.cell(start_row, c, h)
    row_offset = start_row + 1
    for r_idx, row in enumerate(df.itertuples(index=index, name=None), start=row_offset):
        values = row
        if index:
            values = values[1:]
        for c_idx, value in enumerate(values, start=start_col):
            ws.cell(r_idx, c_idx, value)


def build_workbook(
    all_df: pd.DataFrame,
    in_scope: pd.DataFrame,
    exceptions_non_type: pd.DataFrame,
    exceptions_rejected: pd.DataFrame,
    exceptions_feature: pd.DataFrame,
    cycles_df: pd.DataFrame,
    raw_json_path: str,
    output_path: str,
    base_url: str,
    project_key: str = DEFAULT_PROJECT_KEY,
    year: int = DEFAULT_YEAR,
) -> Dict[str, int]:
    wb = Workbook()
    wb.remove(wb.active)

    def _normalize_feature_exceptions(df: pd.DataFrame) -> pd.DataFrame:
        normalized = df.copy()
        normalized.columns = [re.sub(r"[^a-z0-9]+", "_", str(c).strip().lower()).strip("_") for c in normalized.columns]

        alias_candidates = {
            "intent_label": ["intent_label", "intent", "intent_name", "intenttype", "intent_type"],
            "feature_score": ["feature_score", "featureconfidence", "feature_confidence"],
            "bug_score": ["bug_score", "bugconfidence", "bug_confidence"],
            "is_manually_excluded": ["is_manually_excluded", "manual_override", "manually_excluded", "manual_exclusion"],
        }

        for canonical, candidates in alias_candidates.items():
            if canonical in normalized.columns:
                continue
            alias = next((c for c in candidates if c in normalized.columns), None)
            if alias:
                normalized = normalized.rename(columns={alias: canonical})

        missing_intent_label = "intent_label" not in normalized.columns
        required = [
            "key", "summary", "issuetype", "status", "resolution", "priority",
            "created", "resolutiondate", "team", "reporter", "intent_label",
            "feature_score", "bug_score", "is_manually_excluded",
        ]
        for col in required:
            if col not in normalized.columns:
                normalized[col] = None

        if missing_intent_label:
            print("[warn] exceptions_feature missing intent label column; using fallback summary/table structure", file=sys.stderr)

        return normalized

    # Executive Dashboard
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    dashboard_bg = PatternFill("solid", fgColor="FFF6F8FA")
    hero_fill = PatternFill("solid", fgColor="FF153E4D")
    card_fill = PatternFill("solid", fgColor="FFFFFFFF")
    muted_fill = PatternFill("solid", fgColor="FFE9EEF2")
    panel_header_fill = PatternFill("solid", fgColor="FFD9EAF7")
    success_fill = PatternFill("solid", fgColor="FFCFE8D6")
    warning_fill = PatternFill("solid", fgColor="FFFFF2CC")
    danger_fill = PatternFill("solid", fgColor="FFF4CCCC")
    panel_border = Border(
        left=Side(style="thin", color="D8DEE4"),
        right=Side(style="thin", color="D8DEE4"),
        top=Side(style="thin", color="D8DEE4"),
        bottom=Side(style="thin", color="D8DEE4"),
    )
    accent_border = Border(
        left=Side(style="medium", color="153E4D"),
        right=Side(style="thin", color="D8DEE4"),
        top=Side(style="thin", color="D8DEE4"),
        bottom=Side(style="thin", color="D8DEE4"),
    )

    for row in ws.iter_rows(min_row=1, max_row=44, min_col=1, max_col=12):
        for cell in row:
            cell.fill = dashboard_bg

    for col, width in {
        "A": 16, "B": 13, "C": 13, "D": 16, "E": 13, "F": 13,
        "G": 16, "H": 13, "I": 13, "J": 16, "K": 13, "L": 13,
    }.items():
        ws.column_dimensions[col].width = width

    for row_idx, height in {1: 30, 2: 24, 4: 20, 5: 34, 6: 18, 7: 18, 9: 22, 18: 22, 26: 22}.items():
        ws.row_dimensions[row_idx].height = height

    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws["A1"] = "SLA Executive Dashboard"
    ws["A1"].fill = hero_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[1]:
        cell.fill = hero_fill

    ws["A2"] = f"{year} YTD performance command center | Generated {datetime.now().date().isoformat()} | {base_url}"
    ws["A2"].fill = hero_fill
    ws["A2"].font = Font(color="DCE6EA", size=10)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[2]:
        cell.fill = hero_fill

    total_in_scope = len(in_scope)
    total_excluded = len(
        pd.concat(
            [exceptions_non_type, exceptions_rejected, exceptions_feature],
            ignore_index=True
        ).drop_duplicates(subset=["key"])
    )

    frt_population = in_scope[in_scope["frt_pending"] == False]
    res_population = in_scope[in_scope["res_pending"] == False]

    frt_met = int((frt_population["frt_breached"] == False).sum())
    frt_den = len(frt_population)
    frt_pct = percent(frt_met, frt_den) or 0.0

    res_met = int((res_population["res_breached"] == False).sum())
    res_den = len(res_population)
    res_pct = percent(res_met, res_den) or 0.0

    total_breaches = int((in_scope["frt_breached"] == True).sum() + (in_scope["res_breached"] == True).sum())
    both_pct = percent(
        len(in_scope[(in_scope["frt_pending"] == False) & (in_scope["res_pending"] == False) & (in_scope["frt_breached"] == False) & (in_scope["res_breached"] == False)]),
        len(in_scope),
    ) or 0.0

    blended_frt_target = in_scope.apply(lambda r: get_targets(r["priority"], r["quarter"])[0], axis=1).dropna().mean()
    blended_res_target = in_scope.apply(lambda r: get_targets(r["priority"], r["quarter"])[1], axis=1).dropna().mean()

    def _is_number(value) -> bool:
        return isinstance(value, (int, float)) and not pd.isna(value)

    def _format_delta(value, target):
        if not _is_number(value) or not _is_number(target):
            return "No target"
        delta = value - target
        sign = "+" if delta >= 0 else ""
        return f"{sign}{delta:.1%} vs target"

    def _health_fill(value, target):
        if not _is_number(value) or not _is_number(target):
            return muted_fill
        if value >= target:
            return success_fill
        if value >= target - 0.05:
            return warning_fill
        return danger_fill

    def _status_text(frt_value, frt_target, res_value, res_target):
        misses = 0
        for value, target in ((frt_value, frt_target), (res_value, res_target)):
            if _is_number(value) and _is_number(target) and value < target:
                misses += 1
        if misses == 0:
            return "On target"
        if misses == 1:
            return "Watch"
        return "At risk"

    def _style_range(cell_range, fill=card_fill, border=panel_border):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _write_card(cell_range, title, value, subtext, fill):
        ws.merge_cells(cell_range)
        start_cell = ws[cell_range.split(":")[0]]
        start_cell.value = title
        _style_range(cell_range, fill=fill, border=accent_border)
        start_cell.font = Font(color="153E4D", bold=True, size=10)

        start_col = start_cell.column
        start_row = start_cell.row
        value_cell = ws.cell(start_row + 1, start_col)
        subtext_cell = ws.cell(start_row + 2, start_col)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=start_col + 2)
        ws.merge_cells(start_row=start_row + 2, start_column=start_col, end_row=start_row + 3, end_column=start_col + 2)
        value_cell.value = value
        value_cell.font = Font(color="111827", bold=True, size=20)
        value_cell.alignment = Alignment(vertical="center")
        subtext_cell.value = subtext
        subtext_cell.font = Font(color="52616B", size=9)
        subtext_cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + 3, min_col=start_col, max_col=start_col + 2):
            for cell in row:
                cell.fill = fill
                cell.border = accent_border

    _write_card("A4:C4", "FRT Attainment", frt_pct, _format_delta(frt_pct, blended_frt_target), _health_fill(frt_pct, blended_frt_target))
    _write_card("D4:F4", "Resolution Attainment", res_pct, _format_delta(res_pct, blended_res_target), _health_fill(res_pct, blended_res_target))
    _write_card("G4:I4", "Tickets In Scope", total_in_scope, f"{total_excluded} excluded from SLA scope", BLUE_FILL)
    _write_card("J4:L4", "Open Risk", int(((in_scope["res_pending"] == True) & (in_scope["res_breached"] == True)).sum()), f"{total_breaches} total SLA breach flags", warning_fill if total_breaches else success_fill)

    for cell_ref in ("A5", "D5"):
        ws[cell_ref].number_format = "0.0%"

    narrative = narrative_from_data(all_df, in_scope, exceptions_non_type, exceptions_rejected, cycles_df)
    narrative_lines = narrative.splitlines()
    headings = {
        "EXECUTIVE SUMMARY",
        "KEY DRIVERS OF PERFORMANCE",
        "PROCESS WEAKNESSES",
        "RECOMMENDED ACTIONS",
        "HIDDEN INSIGHTS",
    }

    def _section_lines(title):
        collected = []
        in_section = False
        for line in narrative_lines:
            if line == title:
                in_section = True
                continue
            if in_section and line in headings:
                break
            if in_section and line:
                collected.append(line)
        return collected

    def _write_panel(title_cell, body_cell, title, lines, body_range, title_range):
        ws.merge_cells(title_range)
        ws[title_cell] = title
        for row in ws[title_range]:
            for cell in row:
                cell.fill = panel_header_fill
                cell.border = panel_border
        ws[title_cell].font = Font(color="153E4D", bold=True, size=12)
        ws[title_cell].alignment = Alignment(vertical="center")
        ws.merge_cells(body_range)
        ws[body_cell] = "\n".join(lines) if lines else "No data available."
        for row in ws[body_range]:
            for cell in row:
                cell.fill = card_fill
                cell.border = panel_border
        ws[body_cell].font = Font(color="25313B", size=10)
        ws[body_cell].alignment = Alignment(vertical="top", wrap_text=True)

    _write_panel(
        "A9",
        "A10",
        "Executive Summary",
        _section_lines("EXECUTIVE SUMMARY")[:2],
        "A10:F15",
        "A9:F9",
    )

    pr_rows = []
    for p in ["P1 – Critical", "P2 – High", "P3 – Medium", "P4 – Low"]:
        d = in_scope[in_scope["priority"] == p]
        frt_pop = d[d["frt_pending"] == False]
        res_pop = d[d["res_pending"] == False]
        frt_att = float((frt_pop["frt_breached"] == False).mean()) if len(frt_pop) else None
        res_att = float((res_pop["res_breached"] == False).mean()) if len(res_pop) else None
        frt_tgts = d.apply(lambda r: get_targets(r["priority"], r["quarter"])[0], axis=1).dropna()
        res_tgts = d.apply(lambda r: get_targets(r["priority"], r["quarter"])[1], axis=1).dropna()
        frt_tgt = frt_tgts.mean() if len(frt_tgts) else None
        res_tgt = res_tgts.mean() if len(res_tgts) else None
        pr_rows.append([
            p,
            len(d),
            frt_att,
            res_att,
            _status_text(frt_att, frt_tgt, res_att, res_tgt),
        ])
    pr_df = pd.DataFrame(pr_rows, columns=["Priority", "Tickets", "FRT %", "Res %", "Status"])
    ws.merge_cells("H9:L9")
    ws["H9"] = "Priority Health"
    for row in ws["H9:L9"]:
        for cell in row:
            cell.fill = panel_header_fill
            cell.border = panel_border
    ws["H9"].font = Font(color="153E4D", bold=True, size=12)
    write_dataframe(ws, pr_df, start_row=10, start_col=8)
    for cell in ws[10][7:12]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = panel_border
    for row_idx in range(11, 15):
        for col_idx in range(8, 13):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = card_fill
            cell.border = panel_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row_idx, 10).number_format = "0.0%"
        ws.cell(row_idx, 11).number_format = "0.0%"
        status_cell = ws.cell(row_idx, 12)
        status_cell.fill = {
            "On target": success_fill,
            "Watch": warning_fill,
            "At risk": danger_fill,
        }.get(status_cell.value, muted_fill)

    _write_panel(
        "A18",
        "A19",
        "Recommended Actions",
        _section_lines("RECOMMENDED ACTIONS")[:5],
        "A19:F24",
        "A18:F18",
    )

    risk_lines = (
        _section_lines("KEY DRIVERS OF PERFORMANCE")
        + _section_lines("PROCESS WEAKNESSES")
        + _section_lines("HIDDEN INSIGHTS")
    )
    _write_panel(
        "H18",
        "H19",
        "Top Risks",
        risk_lines[:5],
        "H19:L24",
        "H18:L18",
    )

    # Quarterly trend block
    trend = []
    for q in ["Q1", "Q2", "Q3", "Q4"]:
        for p in ["P1 – Critical", "P2 – High", "P3 – Medium", "P4 – Low"]:
            d = in_scope[(in_scope["quarter"] == q) & (in_scope["priority"] == p)]
            pop = d[d["res_pending"] == False]
            trend.append([q, p, float((pop["res_breached"] == False).mean()) if len(pop) else None])
    trend_df = pd.DataFrame(trend, columns=["Quarter", "Priority", "Res %"])
    ws.merge_cells("A26:F26")
    ws["A26"] = "Quarter Trend"
    for row in ws["A26:F26"]:
        for cell in row:
            cell.fill = panel_header_fill
            cell.border = panel_border
    ws["A26"].font = Font(color="153E4D", bold=True, size=12)
    write_dataframe(ws, trend_df, start_row=27, start_col=1)
    for cell in ws[27][0:3]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = panel_border
    for row_idx in range(28, 28 + len(trend_df)):
        for col_idx in range(1, 4):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = card_fill
            cell.border = panel_border
        ws.cell(row_idx, 3).number_format = "0.0%"

    chart = LineChart()
    chart.title = "Resolution attainment by quarter"
    chart.y_axis.title = "SLA Attainment %"
    chart.y_axis.number_format = '0%'
    chart.x_axis.title = "Quarter"
    data = Reference(ws, min_col=3, min_row=27, max_row=27 + len(trend_df))
    cats = Reference(ws, min_col=1, min_row=28, max_row=27 + len(trend_df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "H26")

    def _apply_sheet_title(ws, title: str, end_col: int, subtitle: Optional[str] = None) -> None:
        ws.sheet_view.showGridLines = False
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f"A1:{end_letter}1")
        ws["A1"] = title
        ws["A1"].fill = hero_fill
        ws["A1"].font = Font(color="FFFFFFFF", bold=True, size=16)
        ws["A1"].alignment = Alignment(vertical="center")
        for cell in ws[1]:
            cell.fill = hero_fill
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{end_letter}2")
        ws["A2"] = subtitle or ""
        ws["A2"].fill = muted_fill
        ws["A2"].font = Font(color="52616B", size=9)
        ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        for cell in ws[2]:
            cell.fill = muted_fill
        ws.row_dimensions[2].height = 21

    def _style_header_row(ws, row: int, start_col: int = 1, end_col: Optional[int] = None) -> None:
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = panel_border
        ws.row_dimensions[row].height = 24

    def _style_table_body(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        if end_row < start_row:
            return
        for row_idx in range(start_row, end_row + 1):
            fill = card_fill if row_idx % 2 else PatternFill("solid", fgColor="FFF8FAFC")
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = fill
                cell.border = panel_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _style_section_row(ws, row: int, title: str, end_col: int) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        ws.cell(row, 1, title)
        for col in range(1, end_col + 1):
            cell = ws.cell(row, col)
            cell.fill = panel_header_fill
            cell.border = panel_border
        ws.cell(row, 1).font = Font(color="153E4D", bold=True, size=12)
        ws.cell(row, 1).alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 23

    # SLA Summary
    ws = wb.create_sheet("SLA Summary")
    _apply_sheet_title(ws, "SLA Summary", 14, "Quarterly and YTD SLA attainment by priority and target.")
    start = 3
    blocks = []
    for q in ["Q1", "Q2", "Q3", "Q4"]:
        blocks.append(build_summary_block(in_scope[in_scope["quarter"] == q], q))
    blocks.append(build_summary_block(in_scope, "YTD"))

    delta_headers = []
    for block in blocks:
        _style_section_row(ws, start, str(block.iloc[0]["Quarter"]), 14)
        write_dataframe(ws, block.drop(columns=["Quarter"]), start_row=start + 1, start_col=1)
        _style_header_row(ws, start + 1, 1, len(block.columns) - 1)
        _style_table_body(ws, start + 2, start + 1 + len(block), 1, len(block.columns) - 1)
        for cell in ws[start + 1]:
            if cell.value in ("FRT Delta", "Res Delta"):
                delta_headers.append((start + 1, cell.column, len(block)))
        start += len(block) + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                header = ""
                for scan_row in range(cell.row - 1, 0, -1):
                    value = ws.cell(scan_row, cell.column).value
                    if value:
                        header = str(value)
                        break
                if " %" in header or "Target" in header or "Delta" in header:
                    cell.number_format = "0.0%"

    for header_row, dc, row_count in delta_headers:
        col_letter = get_column_letter(dc)
        start_row_cf = header_row + 1
        end_row_cf = header_row + row_count
        if end_row_cf >= start_row_cf:
            rng = f"{col_letter}{start_row_cf}:{col_letter}{end_row_cf}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
            )
    ws.freeze_panes = "A4"

    # Team Analysis
    ws = wb.create_sheet("Team Analysis")
    _apply_sheet_title(ws, "Team Analysis", 13, "Team-level SLA performance, volume, and risk concentration.")
    team_rows = []
    for team, d in in_scope.groupby("team"):
        frt_pop = d[d["frt_pending"] == False]
        res_pop = d[d["res_pending"] == False]
        frt_att = float((frt_pop["frt_breached"] == False).mean()) if len(frt_pop) else None
        res_att = float((res_pop["res_breached"] == False).mean()) if len(res_pop) else None
        frt_list = d["frt_elapsed_hours"].dropna().tolist()
        res_list = d["res_elapsed_hours"].dropna().tolist()
        blended_targets = d.apply(lambda r: get_targets(r["priority"], r["quarter"])[1], axis=1).dropna()
        blended_res_target = blended_targets.mean() if len(blended_targets) else None
        pri_mix = ", ".join([f"{k}:{v}" for k, v in d["priority"].value_counts().to_dict().items()])
        top_breach_priority = None
        if len(d[(d["res_breached"] == True) | (d["frt_breached"] == True)]):
            top_breach_priority = (
                d[(d["res_breached"] == True) | (d["frt_breached"] == True)]["priority"].value_counts().idxmax()
            )
        team_rows.append(
            {
                "Team": team,
                "Ticket Volume": len(d),
                "Priority Mix": pri_mix,
                "FRT Attainment": frt_att,
                "Res Attainment": res_att,
                "Median FRT": median(frt_list) if frt_list else None,
                "Median Resolution": median(res_list) if res_list else None,
                "P90 Resolution": pctl(res_list, 0.9),
                "Breach Count": int((d["frt_breached"] == True).sum() + (d["res_breached"] == True).sum()),
                "Breach Rate": float(((d["frt_breached"] == True) | (d["res_breached"] == True)).mean()) if len(d) else None,
                "Top Breach Priority": top_breach_priority,
                "Blended Res Target": blended_res_target,
                "Underperform >5pp": (res_att is not None and blended_res_target is not None and (res_att - blended_res_target) < -0.05),
            }
        )
    if team_rows:
        team_df = pd.DataFrame(team_rows).sort_values(["Underperform >5pp", "Res Attainment"], ascending=[False, True])
    else:
        team_df = pd.DataFrame(
            columns=[
                "Team", "Ticket Volume", "Priority Mix", "FRT Attainment", "Res Attainment",
                "Median FRT", "Median Resolution", "P90 Resolution", "Breach Count",
                "Breach Rate", "Top Breach Priority", "Blended Res Target", "Underperform >5pp",
            ]
        )
    write_dataframe(ws, team_df, 3, 1)
    _style_header_row(ws, 3, 1, len(team_df.columns))
    _style_table_body(ws, 4, 3 + len(team_df), 1, len(team_df.columns))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(team_df.columns))}{max(3 + len(team_df), 3)}"

    # Breach Detail
    ws = wb.create_sheet("Breach Detail")
    _apply_sheet_title(ws, "Breach Detail", 16, "Ticket-level SLA breach list with direct Jira links.")
    breach_df = in_scope[(in_scope["frt_breached"] == True) | (in_scope["res_breached"] == True)].copy()
    breach_df = breach_df[[
        "key", "summary", "priority", "issuetype", "team", "assignee", "created", "resolutiondate",
        "frt_elapsed_hours", "frt_goal_hours", "frt_breached",
        "res_elapsed_hours", "res_goal_hours", "res_breached", "quarter", "status", "url"
    ]].rename(columns={
        "key": "Key",
        "summary": "Summary",
        "priority": "Priority",
        "issuetype": "Issue Type",
        "team": "Team",
        "assignee": "Assignee",
        "created": "Created",
        "resolutiondate": "Resolved",
        "frt_elapsed_hours": "FRT elapsed (business hrs)",
        "frt_goal_hours": "FRT goal",
        "frt_breached": "FRT breach Y/N",
        "res_elapsed_hours": "Res elapsed (business hrs)",
        "res_goal_hours": "Res goal",
        "res_breached": "Res breach Y/N",
        "quarter": "Quarter",
        "status": "Status",
        "url": "URL",
    })
    write_dataframe(ws, breach_df, 3, 1)
    _style_header_row(ws, 3, 1, len(breach_df.columns))
    _style_table_body(ws, 4, 3 + len(breach_df), 1, len(breach_df.columns))
    for r in range(4, ws.max_row + 1):
        key_cell = ws.cell(r, 1)
        url = ws.cell(r, breach_df.columns.get_loc("URL") + 1).value
        if url:
            key_cell.hyperlink = url
            key_cell.style = "Hyperlink"
    ws.delete_cols(breach_df.columns.get_loc("URL") + 1)

    for r in range(4, ws.max_row + 1):
        if ws.cell(r, breach_df.columns.get_loc("Res breach Y/N") + 1).value:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = RED_FILL
                ws.cell(r, c).border = panel_border
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{max(ws.max_row, 3)}"

    # Trends & Patterns
    ws = wb.create_sheet("Trends & Patterns")
    _apply_sheet_title(ws, "Trends & Patterns", 10, "Weekly intake, attainment, active-risk aging, and creation heatmap.")
    weekly = in_scope.groupby(["week_start", "priority"]).size().reset_index(name="Tickets")
    write_dataframe(ws, weekly, 3, 1)
    _style_header_row(ws, 3, 1, 3)
    _style_table_body(ws, 4, 3 + len(weekly), 1, 3)

    weekly_frt_source = in_scope[in_scope["frt_pending"] == False]
    if len(weekly_frt_source):
        weekly_frt = (
            weekly_frt_source
            .groupby("week_start")
            .apply(lambda d: float((d["frt_breached"] == False).mean()), include_groups=False)
            .reset_index(name="FRT Attainment")
        )
    else:
        weekly_frt = pd.DataFrame(columns=["week_start", "FRT Attainment"])
    write_dataframe(ws, weekly_frt, 3, 6)
    _style_header_row(ws, 3, 6, 7)
    _style_table_body(ws, 4, 3 + len(weekly_frt), 6, 7)

    weekly_res_source = in_scope[in_scope["res_pending"] == False]
    if len(weekly_res_source):
        weekly_res = (
            weekly_res_source
            .groupby("week_start")
            .apply(lambda d: float((d["res_breached"] == False).mean()), include_groups=False)
            .reset_index(name="Res Attainment")
        )
    else:
        weekly_res = pd.DataFrame(columns=["week_start", "Res Attainment"])
    write_dataframe(ws, weekly_res, 3, 9)
    _style_header_row(ws, 3, 9, 10)
    _style_table_body(ws, 4, 3 + len(weekly_res), 9, 10)

    open_df = in_scope[in_scope["res_pending"] == True].copy()
    age_rows = []
    for _, r in open_df.iterrows():
        if pd.isna(r["res_elapsed_hours"]) or pd.isna(r["res_goal_hours"]) or not r["res_goal_hours"]:
            bucket = "Unknown"
        else:
            ratio = r["res_elapsed_hours"] / r["res_goal_hours"]
            if ratio < 0.25:
                bucket = "<25%"
            elif ratio < 0.75:
                bucket = "25–75%"
            elif ratio <= 1.0:
                bucket = "75–100%"
            else:
                bucket = "Breached"
        age_rows.append({"Priority": r["priority"], "Age Bucket": bucket})
    age_df = pd.DataFrame(age_rows)
    if len(age_df):
        age_pivot = age_df.groupby(["Priority", "Age Bucket"]).size().reset_index(name="Count")
        _style_section_row(ws, 22, "Open Ticket Age", 3)
        write_dataframe(ws, age_pivot, 23, 1)
        _style_header_row(ws, 23, 1, 3)
        _style_table_body(ws, 24, 23 + len(age_pivot), 1, 3)

    heat = in_scope.groupby(["created_day_of_week", "created_hour"]).size().reset_index(name="Count")
    _style_section_row(ws, 22, "Created-Time Heatmap", 10)
    write_dataframe(ws, heat, 23, 6)
    _style_header_row(ws, 23, 6, 8)
    _style_table_body(ws, 24, 23 + len(heat), 6, 8)

    c1 = BarChart()
    c1.title = "Tickets created per week"
    data = Reference(ws, min_col=3, min_row=3, max_row=3 + len(weekly))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(weekly))
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.height = 7
    c1.width = 11
    ws.add_chart(c1, "L3")

    c2 = LineChart()
    c2.title = "FRT attainment rolling view"
    data = Reference(ws, min_col=7, min_row=3, max_row=3 + len(weekly_frt))
    cats = Reference(ws, min_col=6, min_row=4, max_row=3 + len(weekly_frt))
    c2.add_data(data, titles_from_data=True)
    c2.set_categories(cats)
    c2.height = 7
    c2.width = 11
    ws.add_chart(c2, "L18")

    c3 = LineChart()
    c3.title = "Resolution attainment rolling view"
    data = Reference(ws, min_col=10, min_row=3, max_row=3 + len(weekly_res))
    cats = Reference(ws, min_col=9, min_row=4, max_row=3 + len(weekly_res))
    c3.add_data(data, titles_from_data=True)
    c3.set_categories(cats)
    c3.height = 7
    c3.width = 11
    ws.add_chart(c3, "L33")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:C{max(3 + len(weekly), 3)}"

    # Narrative
    ws = wb.create_sheet("Narrative & Recommendations")
    _apply_sheet_title(ws, "Narrative & Recommendations", 1, "Executive narrative, operating risks, and recommended actions.")
    narrative = narrative_from_data(all_df, in_scope, exceptions_non_type, exceptions_rejected, cycles_df)
    for i, line in enumerate(narrative.splitlines(), start=3):
        cell = ws.cell(i, 1, line)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = panel_border
        if line and not line.startswith("- "):
            cell.fill = panel_header_fill
            cell.font = Font(color="153E4D", bold=True, size=12)
            ws.row_dimensions[i].height = 23
        else:
            cell.fill = card_fill
            cell.font = Font(color="25313B", size=10)
            ws.row_dimensions[i].height = 34 if line else 10
    ws.column_dimensions["A"].width = 120

    # Exceptions
    ws = wb.create_sheet("Exceptions")
    _apply_sheet_title(ws, "Exceptions", 14, "Excluded, rejected, and feature-request tickets with summary tables.")
    feature_exceptions = _normalize_feature_exceptions(exceptions_feature)
    _style_section_row(ws, 3, f"Table A - Non-Bug/Issue ticket types ({year} YTD)", 14)
    type_summary = exceptions_non_type["issuetype"].value_counts().reset_index()
    type_summary.columns = ["Issue Type", "Count"]
    write_dataframe(ws, type_summary, 5, 1)
    _style_header_row(ws, 5, 1, 2)
    _style_table_body(ws, 6, 5 + len(type_summary), 1, 2)

    table_a = exceptions_non_type[[
        "key", "summary", "issuetype", "status", "resolution", "priority", "created", "resolutiondate", "team", "reporter"
    ]].rename(columns={
        "key": "Key", "summary": "Summary", "issuetype": "Issue Type", "status": "Status", "resolution": "Resolution",
        "priority": "Priority", "created": "Created", "resolutiondate": "Resolved", "team": "Team", "reporter": "Reporter"
    })
    start_a = 5 + len(type_summary) + 3
    write_dataframe(ws, table_a, start_a, 1)
    _style_header_row(ws, start_a, 1, len(table_a.columns))
    _style_table_body(ws, start_a + 1, start_a + len(table_a), 1, len(table_a.columns))

    feature_keys = set(feature_exceptions["key"].dropna().tolist())
    rejected_for_table_b = exceptions_rejected[~exceptions_rejected["key"].isin(feature_keys)].copy()

    start_b_hdr = start_a + len(table_a) + 4
    _style_section_row(
        ws,
        start_b_hdr,
        f"Table B - Rejected / Won't Do / Duplicate / Cancelled tickets (any issue type, {year} YTD)",
        14,
    )
    res_summary = rejected_for_table_b["resolution"].value_counts().reset_index()
    res_summary.columns = ["Resolution", "Count"]
    write_dataframe(ws, res_summary, start_b_hdr + 2, 1)
    _style_header_row(ws, start_b_hdr + 2, 1, 2)
    _style_table_body(ws, start_b_hdr + 3, start_b_hdr + 2 + len(res_summary), 1, 2)

    matrix = rejected_for_table_b.groupby(["issuetype", "resolution"]).size().reset_index(name="Count")
    write_dataframe(ws, matrix, start_b_hdr + 2, 5)
    _style_header_row(ws, start_b_hdr + 2, 5, 7)
    _style_table_body(ws, start_b_hdr + 3, start_b_hdr + 2 + len(matrix), 5, 7)

    table_b = rejected_for_table_b[[
        "key", "summary", "issuetype", "resolution", "priority", "created", "resolutiondate", "assignee", "reporter"
    ]].copy()
    table_b["Resolution comment"] = None
    table_b = table_b.rename(columns={
        "key": "Key", "summary": "Summary", "issuetype": "Issue Type", "resolution": "Resolution", "priority": "Priority",
        "created": "Created", "resolutiondate": "Resolved", "assignee": "Assignee", "reporter": "Reporter"
    })
    start_b = start_b_hdr + 2 + max(len(res_summary), len(matrix)) + 3
    write_dataframe(ws, table_b, start_b, 1)
    _style_header_row(ws, start_b, 1, len(table_b.columns))
    _style_table_body(ws, start_b + 1, start_b + len(table_b), 1, len(table_b.columns))

    start_c_hdr = start_b + len(table_b) + 4
    _style_section_row(ws, start_c_hdr, f"Table C - Feature requests excluded from SLA ({year} YTD)", 14)

    if len(feature_exceptions):
        feature_summary = (
            feature_exceptions["intent_label"]
            .fillna("Unlabeled")
            .value_counts(dropna=False)
            .rename_axis("Intent Label")
            .reset_index(name="Count")
        )
    else:
        feature_summary = pd.DataFrame([{"Intent Label": "None", "Count": 0}])
    write_dataframe(ws, feature_summary, start_c_hdr + 2, 1)
    _style_header_row(ws, start_c_hdr + 2, 1, 2)
    _style_table_body(ws, start_c_hdr + 3, start_c_hdr + 2 + len(feature_summary), 1, 2)

    table_c = feature_exceptions[[
        "key", "summary", "issuetype", "status", "resolution", "priority",
        "created", "resolutiondate", "team", "reporter", "intent_label",
        "feature_score", "bug_score", "is_manually_excluded"
    ]].rename(columns={
        "key": "Key",
        "summary": "Summary",
        "issuetype": "Issue Type",
        "status": "Status",
        "resolution": "Resolution",
        "priority": "Priority",
        "created": "Created",
        "resolutiondate": "Resolved",
        "team": "Team",
        "reporter": "Reporter",
        "intent_label": "Intent Label",
        "feature_score": "Feature Score",
        "bug_score": "Bug Score",
        "is_manually_excluded": "Manual Override",
    })

    start_c = start_c_hdr + 2 + len(feature_summary) + 3
    write_dataframe(ws, table_c, start_c, 1)
    _style_header_row(ws, start_c, 1, len(table_c.columns))
    _style_table_body(ws, start_c + 1, start_c + len(table_c), 1, len(table_c.columns))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{start_a}:{get_column_letter(len(table_a.columns))}{max(start_a + len(table_a), start_a)}"

    # Inclusion Logic
    ws = wb.create_sheet("Inclusion Logic")
    _apply_sheet_title(
        ws,
        "Inclusion Logic",
        13,
        "All pulled Jira tickets with the inclusion/exclusion decision and the rule flags behind it.",
    )
    logic_columns = [
        "key",
        "scope_decision",
        "scope_reason",
        "issuetype",
        "resolution",
        "status",
        "priority",
        "summary",
        "is_in_scope_issue_type",
        "is_not_accepted",
        "is_disguised_feature_request",
        "is_manual_bug_include",
        "url",
    ]
    logic_df = all_df.copy()
    for col in logic_columns:
        if col not in logic_df.columns:
            logic_df[col] = None
    logic_df = logic_df[logic_columns].rename(
        columns={
            "key": "Key",
            "scope_decision": "Scope Decision",
            "scope_reason": "Scope Reason",
            "issuetype": "Issue Type",
            "resolution": "Resolution",
            "status": "Status",
            "priority": "Priority",
            "summary": "Summary",
            "is_in_scope_issue_type": "In-Scope Type",
            "is_not_accepted": "Excluded Resolution",
            "is_disguised_feature_request": "Feature Request Rule",
            "is_manual_bug_include": "Manual Include",
            "url": "URL",
        }
    )
    write_dataframe(ws, logic_df, 3, 1)
    _style_header_row(ws, 3, 1, len(logic_df.columns))
    _style_table_body(ws, 4, 3 + len(logic_df), 1, len(logic_df.columns))
    for row_idx in range(4, 4 + len(logic_df)):
        key_cell = ws.cell(row_idx, 1)
        url = ws.cell(row_idx, logic_df.columns.get_loc("URL") + 1).value
        if url:
            key_cell.hyperlink = url
            key_cell.style = "Hyperlink"
        decision_cell = ws.cell(row_idx, logic_df.columns.get_loc("Scope Decision") + 1)
        if decision_cell.value == "Included":
            decision_cell.fill = success_fill
        else:
            decision_cell.fill = warning_fill
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(logic_df.columns))}{max(3 + len(logic_df), 3)}"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["H"].width = 56
    ws.column_dimensions["M"].width = 46

    # Methodology
    ws = wb.create_sheet("Methodology")
    _apply_sheet_title(ws, "Methodology", 2, "Source assumptions, scope logic, endpoints, and extract metadata.")
    meth = [
        [
            "Exact JQL used for raw extract",
            f'project = {project_key} AND created >= "{year}-01-01" AND created < "{year + 1}-01-01" ORDER BY created ASC',
        ],
        ["In-scope issue types", "; ".join(sorted(IN_SCOPE_ISSUE_TYPES))],
        ["Accepted logic", f"resolution is empty OR resolution not in {sorted(NOT_ACCEPTED_RESOLUTIONS)}"],
        ["Rejected mapping", ", ".join(sorted(NOT_ACCEPTED_RESOLUTIONS))],
        ["SLA endpoint used", "/rest/servicedeskapi/request/{issueIdOrKey}/sla"],
        ["FRT SLA names read", ", ".join(sorted(SLA_FRT_NAMES))],
        ["Resolution SLA names read", ", ".join(sorted(SLA_RES_NAMES))],
        ["Team field used", f"{TEAM_FIELD_NAME} ({TEAM_FIELD_ID})"],
        ["Q1 target assumption", "Same as Q2"],
        ["Quarter boundaries used", "Q1 Jan-Mar; Q2 Apr-Jun; Q3 Jul-Sep; Q4 Oct-Dec"],
        ["Raw JSON sidecar", raw_json_path],
        ["Known limitations", "This workbook uses true SLA cycles from JSM. Resolution comments are not extracted. Priority downgrade evidence is simplified from changelog history."],
        ["Extraction timestamp UTC", datetime.now(timezone.utc).isoformat()],
        ["Ticket count", len(all_df)],
        ["In-scope count", len(in_scope)],
        ["Exceptions non type", len(exceptions_non_type)],
        ["Exceptions rejected", len(exceptions_rejected)],
        ["Feature-request exclusions", len(exceptions_feature)],
    ]
    for i, row in enumerate(meth, start=3):
        ws.cell(i, 1, row[0]).font = BOLD
        ws.cell(i, 2, row[1])
        for c in range(1, 3):
            cell = ws.cell(i, c)
            cell.fill = card_fill if i % 2 else PatternFill("solid", fgColor="FFF8FAFC")
            cell.border = panel_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120

    # Global formatting
    for sheet in wb.worksheets:
        autosize(sheet)

    def format_percent_columns(ws):
        percent_ranges = []
        for row in ws.iter_rows():
            for cell in row:
                header = str(cell.value)
                if any(token in header for token in ("%", "Target", "Delta", "Rate", "Attainment")):
                    percent_ranges.append((cell.column, cell.row + 1))
        for col_idx, min_row in percent_ranges:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=min_row, max_row=ws.max_row).__next__():
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"

    format_percent_columns(wb["SLA Summary"])
    format_percent_columns(wb["Team Analysis"])
    format_percent_columns(wb["Executive Dashboard"])

    # Team Analysis highlighting
    team_ws = wb["Team Analysis"]
    headers = {team_ws.cell(3, c).value: c for c in range(1, team_ws.max_column + 1)}
    under_col = headers.get("Underperform >5pp")
    if under_col:
        for r in range(4, team_ws.max_row + 1):
            if team_ws.cell(r, under_col).value is True:
                for c in range(1, team_ws.max_column + 1):
                    team_ws.cell(r, c).fill = AMBER_FILL

    wb.save(output_path)

    return {
        "pulled": len(all_df),
        "in_scope": len(in_scope),
        "exceptions_total": len(
            pd.concat(
                [exceptions_non_type, exceptions_rejected, exceptions_feature],
                ignore_index=True
            ).drop_duplicates(subset=["key"])
        ),
        "exceptions_non_type": len(exceptions_non_type),
        "exceptions_rejected": len(exceptions_rejected),
        "exceptions_feature": len(exceptions_feature),
    }
