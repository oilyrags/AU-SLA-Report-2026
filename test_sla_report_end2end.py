#!/usr/bin/env python3
import importlib.util
import pathlib
import sys
import tempfile
import unittest

import pandas as pd
from openpyxl import load_workbook


MODULE_PATH = pathlib.Path(__file__).with_name("sla_report_end2end.py")
SPEC = importlib.util.spec_from_file_location("sla_report_end2end", MODULE_PATH)
sla_report_end2end = importlib.util.module_from_spec(SPEC)
assert SPEC is not None and SPEC.loader is not None
sys.modules[SPEC.name] = sla_report_end2end
SPEC.loader.exec_module(sla_report_end2end)

import report_publish


class ClassifyScopeBehaviorTests(unittest.TestCase):
    def test_disguised_feature_requests_and_email_requests_scope(self):
        df = pd.DataFrame(
            [
                {
                    "key": "ASD-1992",
                    "issuetype": "Technical support",
                    "resolution": "Resolved",
                    "summary": "Bulk Edit Vehicles in Retool (Merbag Issue)",
                    "labels": "ServiceDesk",
                    "team": "Inventory - Retool",
                    "domain": "Inventory - Retool",
                    "components": "",
                },
                {
                    "key": "ASD-1993",
                    "issuetype": "Technical support",
                    "resolution": "Resolved",
                    "summary": "Searchfunction in Retool by Wagennummer (Sellervehicleid)",
                    "labels": "ServiceDesk",
                    "team": "Inventory - Retool",
                    "domain": "Inventory - Retool",
                    "components": "",
                },
                {
                    "key": "ASD-1999",
                    "issuetype": "Email request",
                    "resolution": "Resolved",
                    "summary": "63636 - Check/Add Import report E-Mail",
                    "labels": "ServiceDesk",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                },
                {
                    "key": "ASD-2000",
                    "issuetype": "Email request",
                    "resolution": "Resolved",
                    "summary": "BMW / Mini Qualilogo text changes",
                    "labels": "ServiceDesk",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                },
                {
                    "key": "ASD-2189",
                    "issuetype": "Email request",
                    "resolution": None,
                    "summary": "Impementation of the Ombudsstelle-Link in AS24/MS24 homepage-footer",
                    "labels": "ServiceDesk",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                },
                {
                    "key": "ASD-3001",
                    "issuetype": "Email request",
                    "resolution": "Rejected",
                    "summary": "Rejected email request",
                    "labels": "ServiceDesk",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                },
                {
                    "key": "ASD-3002",
                    "issuetype": "Email request",
                    "resolution": "Resolved",
                    "summary": "Please create bulk edit flow in Retool",
                    "labels": "ServiceDesk",
                    "team": "Inventory - Retool",
                    "domain": "Inventory - Retool",
                    "components": "",
                },
                {
                    "key": "ASD-3003",
                    "issuetype": "emal request",
                    "resolution": "Resolved",
                    "summary": "Customer reported upload failure by email",
                    "labels": "ServiceDesk",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                },
            ]
        )

        all_df, in_scope, _, exceptions_rejected, exceptions_feature = sla_report_end2end.classify_scope(df)

        in_scope_keys = set(in_scope["key"].tolist())
        rejected_exception_keys = set(exceptions_rejected["key"].tolist())
        feature_exception_keys = set(exceptions_feature["key"].tolist())

        self.assertIn("ASD-1992", feature_exception_keys)
        self.assertIn("ASD-1993", feature_exception_keys)
        self.assertIn("ASD-3002", feature_exception_keys)
        self.assertNotIn("ASD-1992", in_scope_keys)
        self.assertNotIn("ASD-1993", in_scope_keys)
        self.assertNotIn("ASD-3002", in_scope_keys)

        self.assertIn("ASD-1999", in_scope_keys)
        self.assertIn("ASD-2000", in_scope_keys)
        self.assertIn("ASD-2189", in_scope_keys)
        self.assertIn("ASD-3003", in_scope_keys)
        self.assertIn("ASD-3001", rejected_exception_keys)
        self.assertNotIn("ASD-3001", in_scope_keys)

        by_key = all_df.set_index("key")
        self.assertTrue(bool(by_key.loc["ASD-1992", "is_disguised_feature_request"]))
        self.assertTrue(bool(by_key.loc["ASD-1993", "is_disguised_feature_request"]))
        self.assertTrue(bool(by_key.loc["ASD-1999", "is_manual_bug_include"]))
        self.assertTrue(bool(by_key.loc["ASD-2000", "is_manual_bug_include"]))
        self.assertFalse(bool(by_key.loc["ASD-2189", "is_manual_bug_include"]))
        self.assertTrue(bool(by_key.loc["ASD-2189", "is_bug_issue_type"]))

        self.assertEqual(by_key.loc["ASD-2189", "scope_decision"], "Included")
        self.assertIn("issue type is in scope", by_key.loc["ASD-2189", "scope_reason"])
        self.assertEqual(by_key.loc["ASD-3001", "scope_decision"], "Excluded")
        self.assertIn("resolution", by_key.loc["ASD-3001", "scope_reason"])
        self.assertEqual(by_key.loc["ASD-3002", "scope_decision"], "Excluded")
        self.assertIn("feature request", by_key.loc["ASD-3002", "scope_reason"])
        self.assertEqual(by_key.loc["ASD-3003", "scope_decision"], "Included")


class DashboardPayloadTests(unittest.TestCase):
    def test_build_dashboard_payload_reports_tickets_in_scope_and_open_risk(self):
        in_scope = pd.DataFrame(
            [
                {
                    "frt_has_sla": True,
                    "frt_pending": False,
                    "frt_breached": False,
                    "res_has_sla": True,
                    "res_pending": False,
                    "res_breached": False,
                },
                {
                    "frt_has_sla": True,
                    "frt_pending": False,
                    "frt_breached": True,
                    "res_has_sla": True,
                    "res_pending": False,
                    "res_breached": False,
                },
                {
                    "frt_has_sla": False,
                    "frt_pending": False,
                    "frt_breached": False,
                    "res_has_sla": False,
                    "res_pending": True,
                    "res_breached": True,
                },
            ]
        )

        payload = report_publish.build_dashboard_payload(in_scope)
        metrics = payload.set_index("metric")["value"].to_dict()

        self.assertEqual(metrics["Tickets In Scope"], 3)
        self.assertEqual(metrics["Open Risk"], 1)
        self.assertIn("FRT Attainment %", metrics)
        self.assertIn("Resolution Attainment %", metrics)

    def test_sla_report_end2end_reexports_dashboard_helper(self):
        self.assertIs(
            sla_report_end2end.build_dashboard_payload,
            report_publish.build_dashboard_payload,
        )


class EmptyReportTests(unittest.TestCase):
    def test_workbook_handles_report_with_no_in_scope_tickets(self):
        all_df = pd.DataFrame(
            [
                {
                    "key": "ASD-3000",
                    "summary": "Feature request outside SLA scope",
                    "issuetype": "Change request",
                    "status": "Closed",
                    "resolution": "Resolved",
                    "priority": "P4 – Low",
                    "created": "2026-04-24T12:00:00.000+0000",
                    "resolutiondate": "2026-04-24T13:00:00.000+0000",
                    "team": "Shared Service",
                    "domain": "Shared Service",
                    "components": "",
                    "labels": "",
                    "reporter": "user@example.com",
                    "assignee": "agent@example.com",
                    "quarter": "Q2",
                    "week_start": "2026-04-20",
                    "created_day_of_week": "Friday",
                    "created_hour": 12,
                    "url": "https://example.atlassian.net/browse/ASD-3000",
                    "frt_has_sla": False,
                    "frt_pending": False,
                    "frt_breached": False,
                    "frt_elapsed_hours": None,
                    "frt_goal_hours": None,
                    "frt_total_cycles": 0,
                    "res_has_sla": False,
                    "res_pending": False,
                    "res_breached": False,
                    "res_elapsed_hours": None,
                    "res_goal_hours": None,
                    "res_total_cycles": 0,
                    "business_to_wall_clock_ratio": None,
                    "priority_downgrade_evidence": False,
                }
            ]
        )
        all_df, in_scope, non_type, rejected, feature = sla_report_end2end.classify_scope(all_df)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            sla_report_end2end.build_workbook(
                all_df=all_df,
                in_scope=in_scope,
                exceptions_non_type=non_type,
                exceptions_rejected=rejected,
                exceptions_feature=feature,
                cycles_df=pd.DataFrame(),
                raw_json_path="/tmp/raw.json",
                output_path=output_path,
                base_url="https://example.atlassian.net",
            )

            workbook = load_workbook(output_path)
            narrative_values = [
                cell
                for row in workbook["Narrative & Recommendations"].iter_rows(values_only=True)
                for cell in row
                if cell is not None
            ]
            self.assertIn("No in-scope tickets were found for this report period.", narrative_values)
        finally:
            pathlib.Path(output_path).unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
